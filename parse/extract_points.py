import csv
import io
import re
import sys
import warnings

warnings.filterwarnings("ignore")  # silence openpyxl extension warnings

try:
    import msoffcrypto
    import openpyxl
except ImportError:
    sys.exit("Missing deps. Run:  pip install msoffcrypto-tool openpyxl")

# ---------------------------------------------------------------------------
# SETTINGS -- edit these instead of passing command-line arguments
# ---------------------------------------------------------------------------
INPUT_FILE = "AI TEST - DDC 3 IRQAH.xlsx"
PASSWORD = "Estimation"
SHEET = "Sheet1"
OUTPUT_FILE = "test_points_tags.csv"
# ---------------------------------------------------------------------------

RED = "FFFF0000"          # font color that marks tags (col C) and meta (col D)
DDC_RE = re.compile(r"^DDC\s*\d+", re.I)

# Column E (5) holds QTY -- the point's multiplicity, used only to validate
# the Total columns below; it never determines I/O type.
QTY_COL = 5

# Columns F..I (6..9) hold the physical I/O type flags: DI, DO, AI, AO. This is
# the AUTHORITATIVE source of a point's I/O type -- see classify_io_type() in
# main.py. A genuine physical point carries a numeric one-hot here (e.g.
# 0 1 0 0 / 1 0 0 0). Rows with these columns empty are NOT physical points --
# they are integration/software sub-points (e.g. the EV-charging alarms listed
# under a BACnet device) or header rows, and must not be emitted.
IO_TYPE_COLS = (6, 7, 8, 9)

# Columns J..M (10..13) hold the "Total Points" sub-table: the totalized
# DI/BI, DO/BO, AI/UI, AO/UO counts that sit to the right of the physical
# I/O flags. This is a VALIDATION total (expected to equal QTY * the matching
# F..I flag), never the type source. These are copied through verbatim: a
# numeric cell (including a meaningful 0) is kept as-is; a blank / non-numeric
# cell becomes "None".
TOTAL_POINTS_COLS = (10, 11, 12, 13)

# Column A (Item Code) prefixes that mark summary / footer rows in the template.
# These are the workbook's own machine codes -- present in the master file even
# though the visible label may vary. Matched case-insensitively by prefix.
SUMMARY_CODE_PREFIXES = ("DDCNR", "DDCTOT", "CONTR", "DDCNF")

# Fallback for sheets where column A is blank (e.g. AI-extracted single-DDC tabs):
# the template's fixed footer labels in column D. Matched by exact, normalized,
# case-insensitive equality -- extend this set if new footer rows appear.
SUMMARY_LABELS = {
    "spare physical points",
    "ddc total physical points",
    "ddc total integration devices / points",
}


def _norm(s):
    """Lower-case and collapse whitespace for stable label comparison."""
    return re.sub(r"\s+", " ", s).strip().lower() if isinstance(s, str) else s


def is_summary_row(a_val, d_val):
    """True if the row is a summary/footer row (not a real point)."""
    if isinstance(a_val, str):
        code = a_val.strip().upper()
        if any(code.startswith(p) for p in SUMMARY_CODE_PREFIXES):
            return True
    return _norm(d_val) in SUMMARY_LABELS


def load_sheet(path, password, sheet, data_only):
    """Return one worksheet, decrypting in memory only if the file is encrypted."""
    with open(path, "rb") as fh:
        off = msoffcrypto.OfficeFile(fh)
        if off.is_encrypted():
            dec = io.BytesIO()
            off.load_key(password=password)
            off.decrypt(dec)
            source = dec
        else:
            fh.seek(0)
            source = io.BytesIO(fh.read())
    wb = openpyxl.load_workbook(source, data_only=data_only)
    if sheet not in wb.sheetnames:
        sys.exit(f'Sheet "{sheet}" not found. Available: {wb.sheetnames}')
    return wb[sheet]


def font_rgb(cell):
    f = cell.font
    return f.color.rgb if (f.color and f.color.type == "rgb") else None


def extract(path, password, sheet):
    values = load_sheet(path, password, sheet, data_only=True)   # computed text
    styles = load_sheet(path, password, sheet, data_only=False)  # font colors

    def text(r, col):
        v = values.cell(r, col).value
        return v.strip() if isinstance(v, str) else v

    def numeric_cols(r, cols):
        """Read a set of columns as-is: numbers (incl. 0) kept, blanks -> "None"."""
        out = []
        for c in cols:
            v = values.cell(r, c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                out.append(v)
            else:
                out.append("None")
        return out

    def qty(r):
        v = values.cell(r, QTY_COL).value
        return v if isinstance(v, (int, float)) and not isinstance(v, bool) else "None"

    def is_physical_point(r):
        """True only if the row carries a numeric DI/DO/AI/AO count (>=1)."""
        seen, total = False, 0
        for c in IO_TYPE_COLS:
            v = values.cell(r, c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                seen, total = True, total + v
        return seen and total >= 1

    rows = []
    ddc = etype = etag = None
    for r in range(1, values.max_row + 1):
        aval = text(r, 1)
        cval = text(r, 3)
        ccol = font_rgb(styles.cell(r, 3))
        dval = text(r, 4)
        dcol = font_rgb(styles.cell(r, 4))

        # table header resets everything
        if isinstance(cval, str) and DDC_RE.match(cval):
            ddc, etype, etag = cval, None, None
            continue

        # summary / footer rows (Spare / DDC Total ...) are not points
        if is_summary_row(aval, dval):
            continue

        # column C: equipment type (black) or tag (red)
        if isinstance(cval, str) and cval and cval != "Device":
            if ccol == RED:
                etag = cval
            else:
                etype, etag = cval, None

        # column D: a real point (non-red, has a current equipment type) AND it
        # must carry a physical DI/DO/AI/AO count -- otherwise it is an
        # integration/software sub-point, not a physical point.
        if isinstance(dval, str) and dval and dval != "Point's Tag" and dcol != RED:
            if etype and is_physical_point(r):
                combined = " ".join(x for x in (etype, etag, dval) if x)
                rows.append([ddc or "", etype, etag or "", dval, combined, qty(r)]
                            + numeric_cols(r, IO_TYPE_COLS)
                            + numeric_cols(r, TOTAL_POINTS_COLS))
    return rows


def main():
    rows = extract(INPUT_FILE, PASSWORD, SHEET)

    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["DDC", "Equipment Type", "Equipment Tag", "Point Tag", "Full Combined", "QTY",
                    "DI", "DO", "AI", "AO",
                    "Total_DI", "Total_DO", "Total_AI", "Total_AO"])
        w.writerows(rows)

    ddcs = len({r[0] for r in rows})
    print(f"Wrote {len(rows)} point rows across {ddcs} DDCs -> {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
